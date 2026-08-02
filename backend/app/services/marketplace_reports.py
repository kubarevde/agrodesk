"""Marketplace showcase orders report — isolated from farm /api/reports.

Honest metrics only: заявка counts and «сумма заявок» estimated from current
listing prices (no payment / revenue claims). Orders have no price snapshot.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID

from openpyxl.workbook.workbook import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.marketplace import MarketListing, MarketOrder, MarketSellerProfile
from app.services.excel_styles import new_workbook, write_table
from app.services.marketplace_seller import require_marketplace_enabled

ORDER_STATUS_LABELS_RU: dict[str, str] = {
    'new': 'Новая',
    'contacted': 'Связались',
    'confirmed': 'Подтверждена',
    'completed': 'Выполнена',
    'cancelled': 'Отменена',
}

ORDER_STATUSES: tuple[str, ...] = (
    'new',
    'contacted',
    'confirmed',
    'completed',
    'cancelled',
)


def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _fmt_date(value: date) -> str:
    return value.strftime('%d.%m.%Y')


def _fmt_datetime(value: datetime | None) -> str:
    if value is None:
        return ''
    return value.strftime('%d.%m.%Y %H:%M')


@dataclass(frozen=True)
class MarketOrderReportRow:
    order_id: UUID
    created_at: datetime
    updated_at: datetime
    status: str
    listing_id: UUID
    listing_title: str
    listing_unit: str
    listing_price: Decimal
    quantity: Decimal
    estimated_amount: Decimal
    buyer_name: str
    buyer_phone: str
    buyer_comment: str | None
    seller_display_name: str


@dataclass(frozen=True)
class StatusBucket:
    status: str
    label: str
    orders_count: int
    quantity_sum: Decimal
    estimated_amount_sum: Decimal


@dataclass(frozen=True)
class MarketOrdersReport:
    from_date: date
    to_date: date
    org_id: UUID
    seller_display_name: str
    orders_count: int
    quantity_sum: Decimal
    estimated_amount_sum: Decimal
    status_breakdown: list[StatusBucket]
    rows: list[MarketOrderReportRow]
    amount_disclaimer: str


AMOUNT_DISCLAIMER = (
    'Сумма заявок — оценка по текущим ценам объявлений, не выручка и не оплата. '
    'Заявка витрины не является продажей со склада КФХ.'
)


async def fetch_market_orders_report(
    db: AsyncSession,
    org_id: UUID,
    from_date: date,
    to_date: date,
    *,
    status_filter: str | None = None,
) -> MarketOrdersReport:
    await require_marketplace_enabled(db, org_id)

    if to_date < from_date:
        from_date, to_date = to_date, from_date

    seller = await db.scalar(
        select(MarketSellerProfile).where(MarketSellerProfile.org_id == org_id)
    )
    seller_name = seller.display_name if seller is not None else ''

    created_day = func.date(MarketOrder.created_at)
    filters = [
        MarketListing.org_id == org_id,
        created_day >= from_date,
        created_day <= to_date,
    ]
    if status_filter:
        filters.append(MarketOrder.status == status_filter)

    result = await db.execute(
        select(
            MarketOrder,
            MarketListing,
            MarketSellerProfile.display_name,
        )
        .join(MarketListing, MarketListing.id == MarketOrder.listing_id)
        .join(
            MarketSellerProfile,
            MarketSellerProfile.id == MarketListing.seller_profile_id,
        )
        .where(*filters)
        .order_by(MarketOrder.created_at.desc())
    )

    rows: list[MarketOrderReportRow] = []
    qty_total = Decimal('0')
    amount_total = Decimal('0')
    count_by: dict[str, int] = {key: 0 for key in ORDER_STATUSES}
    qty_by: dict[str, Decimal] = {key: Decimal('0') for key in ORDER_STATUSES}
    amount_by: dict[str, Decimal] = {key: Decimal('0') for key in ORDER_STATUSES}

    for order, listing, profile_name in result.all():
        qty = _to_decimal(order.quantity)
        price = _to_decimal(listing.price)
        amount = (qty * price).quantize(Decimal('0.01'))
        rows.append(
            MarketOrderReportRow(
                order_id=order.id,
                created_at=order.created_at,
                updated_at=order.updated_at,
                status=order.status,
                listing_id=listing.id,
                listing_title=listing.title,
                listing_unit=listing.unit,
                listing_price=price,
                quantity=qty,
                estimated_amount=amount,
                buyer_name=order.buyer_name,
                buyer_phone=order.buyer_phone,
                buyer_comment=order.buyer_comment,
                seller_display_name=str(profile_name or seller_name),
            )
        )
        qty_total += qty
        amount_total += amount
        if order.status in count_by:
            count_by[order.status] += 1
            qty_by[order.status] += qty
            amount_by[order.status] += amount

    breakdown = [
        StatusBucket(
            status=key,
            label=ORDER_STATUS_LABELS_RU.get(key, key),
            orders_count=count_by[key],
            quantity_sum=qty_by[key],
            estimated_amount_sum=amount_by[key],
        )
        for key in ORDER_STATUSES
    ]

    return MarketOrdersReport(
        from_date=from_date,
        to_date=to_date,
        org_id=org_id,
        seller_display_name=seller_name,
        orders_count=len(rows),
        quantity_sum=qty_total,
        estimated_amount_sum=amount_total,
        status_breakdown=breakdown,
        rows=rows,
        amount_disclaimer=AMOUNT_DISCLAIMER,
    )


async def build_market_orders_workbook(
    db: AsyncSession,
    org_id: UUID,
    from_date: date,
    to_date: date,
    *,
    status_filter: str | None = None,
) -> Workbook:
    report = await fetch_market_orders_report(
        db,
        org_id,
        from_date,
        to_date,
        status_filter=status_filter,
    )

    workbook = new_workbook()
    ws = workbook.active
    ws.title = 'Заявки витрины'

    detail_rows: list[list[object]] = []
    for row in report.rows:
        detail_rows.append(
            [
                _fmt_datetime(row.created_at),
                ORDER_STATUS_LABELS_RU.get(row.status, row.status),
                row.listing_title,
                float(row.listing_price),
                row.listing_unit,
                float(row.quantity),
                float(row.estimated_amount),
                row.buyer_name,
                row.buyer_phone,
                row.buyer_comment or '',
                row.seller_display_name,
                _fmt_datetime(row.updated_at),
            ]
        )

    write_table(
        ws,
        [
            'Дата заявки',
            'Статус',
            'Объявление',
            'Цена объявления',
            'Ед.',
            'Кол-во',
            'Оценка суммы',
            'Покупатель',
            'Телефон',
            'Комментарий',
            'Продавец',
            'Обновлено',
        ],
        detail_rows,
        [
            'ИТОГО',
            f'{report.orders_count} заявок',
            '',
            '',
            '',
            float(report.quantity_sum),
            float(report.estimated_amount_sum),
            '',
            '',
            '',
            '',
            '',
        ],
    )

    # Title / disclaimer above table would shift write_table — append note below.
    note_row = ws.max_row + 2
    ws.cell(
        row=note_row,
        column=1,
        value=(
            f'Период: {_fmt_date(report.from_date)} — {_fmt_date(report.to_date)}. '
            f'{AMOUNT_DISCLAIMER}'
        ),
    )

    summary = workbook.create_sheet('Сводка по статусам')
    summary_rows = [
        [
            bucket.label,
            bucket.orders_count,
            float(bucket.quantity_sum),
            float(bucket.estimated_amount_sum),
        ]
        for bucket in report.status_breakdown
    ]
    write_table(
        summary,
        [
            'Статус',
            'Заявок',
            'Кол-во',
            'Сумма заявок (оценка)',
        ],
        summary_rows,
        [
            'Всего',
            report.orders_count,
            float(report.quantity_sum),
            float(report.estimated_amount_sum),
        ],
    )
    summary.cell(
        row=summary.max_row + 2,
        column=1,
        value=AMOUNT_DISCLAIMER,
    )

    return workbook
