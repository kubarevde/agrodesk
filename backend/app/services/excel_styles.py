from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
EVEN_ROW_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)


def style_header_row(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def style_even_row(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = EVEN_ROW_FILL


def style_total_row(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL


def autosize_columns(ws: Worksheet, max_col: int, min_width: int = 10, max_width: int = 40) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        values = [str(ws.cell(row=row, column=col).value or '') for row in range(1, ws.max_row + 1)]
        width = min(max(len(value) for value in values) + 2, max_width)
        ws.column_dimensions[letter].width = max(width, min_width)


def write_table(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[object]],
    totals_row: list[object] | None = None,
) -> None:
    max_col = len(headers)
    ws.append(headers)
    style_header_row(ws, 1, max_col)

    for index, row in enumerate(rows, start=2):
        ws.append(row)
        if index % 2 == 0:
            style_even_row(ws, index, max_col)

    if totals_row is not None:
        total_row_num = len(rows) + 2
        ws.append(totals_row)
        style_total_row(ws, total_row_num, max_col)

    autosize_columns(ws, max_col)


def new_workbook() -> Workbook:
    from openpyxl import Workbook

    return Workbook()
