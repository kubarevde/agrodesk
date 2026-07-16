import calendar
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from fastapi.responses import StreamingResponse
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.agro_plan import AgroPlan
from app.models.dictionary import OrgDictionary
from app.models.employee import Employee
from app.models.equipment_log import EquipmentMaintenance, EquipmentMeterLog
from app.models.expense import Expense
from app.models.implement import Implement, ImplementMaintenance
from app.models.inventory import InventoryItem, InventoryOperation
from app.models.reference import Equipment, Location
from app.models.shift import Shift, ShiftStatus
from app.models.shipment import Shipment
from app.services.equipment_meters import calc_meter_label
from app.services.excel_styles import new_workbook, write_table
from app.services.shifts import calc_duration_from_datetimes, combine_date_time

EXPENSE_CATEGORY_LABELS = {
    # Legacy compatibility for historical expense.category codes before custom dictionaries.
    'fuel': 'Топливо',
    'fertilizer': 'Удобрения',
    'parts': 'Запчасти',
    'salary': 'Зарплата',
    'rent': 'Аренда',
    'other': 'Прочее',
}

INVENTORY_CATEGORY_LABELS = {
    # Legacy compatibility for historical inventory.category codes.
    'fuel': 'Топливо',
    'fertilizer': 'Удобрения',
    'parts': 'Запчасти',
    'seeds': 'Семена',
    'chemicals': 'СЗР',
    'other': 'Прочее',
}


async def load_dictionary_labels(
    db: AsyncSession,
    org_id: UUID | None,
    dict_type: str,
    fallback: dict[str, str] | None = None,
    *,
    include_inactive: bool = True,
) -> dict[str, str]:
    """Map dictionary code (and name) → display name for Excel / reports.

    Includes inactive rows by default so historical records keep their labels
    after a category was deactivated. Active names win over inactive on conflict.
    """
    labels = dict(fallback or {})
    if org_id is None:
        return labels
    query = select(OrgDictionary).where(
        OrgDictionary.org_id == org_id,
        OrgDictionary.type == dict_type,
    )
    if not include_inactive:
        query = query.where(OrgDictionary.is_active.is_(True))
    result = await db.execute(query.order_by(OrgDictionary.is_active.asc()))
    for row in result.scalars().all():
        labels[row.code] = row.name
        labels[row.name] = row.name
    return labels


def resolve_label(value: str | None, labels: dict[str, str]) -> str:
    if not value:
        return ''
    return labels.get(value, value)

OPERATION_TYPE_LABELS = {
    'income': 'Приход',
    'expense': 'Расход',
}

SHIFT_STATUS_LABELS = {
    'open': 'Открыта',
    'closed': 'Закрыта',
}

IMPLEMENT_CONDITION_LABELS = {
    'good': 'Хорошее',
    'fair': 'Удовлетворительное',
    'poor': 'Плохое',
    'repair': 'В ремонте',
}

AGRO_STATUS_LABELS = {
    'planned': 'Запланировано',
    'in_progress': 'В работе',
    'done': 'Выполнено',
    'cancelled': 'Отменено',
}

MONTH_LABELS = [
    'Янв',
    'Фев',
    'Мар',
    'Апр',
    'Май',
    'Июн',
    'Июл',
    'Авг',
    'Сен',
    'Окт',
    'Ноя',
    'Дек',
]


def parse_month(month: str) -> tuple[date, date]:
    year, mon = map(int, month.split('-'))
    last_day = calendar.monthrange(year, mon)[1]
    return date(year, mon, 1), date(year, mon, last_day)


def fmt_date(value: date) -> str:
    return value.strftime('%d.%m.%Y')


def fmt_time(value) -> str:
    if value is None:
        return ''
    return value.strftime('%H:%M')


def shift_hours(shift: Shift) -> float:
    if shift.duration_rounded is not None:
        return float(shift.duration_rounded)
    if shift.status == ShiftStatus.open:
        now = datetime.now()
        start_dt = combine_date_time(shift.date, shift.start_time)
        minutes = calc_duration_from_datetimes(start_dt, now)
        return round(minutes / 60, 2)
    return 0.0


def to_number(value: Decimal | float | int | None) -> float:
    if value is None:
        return 0.0
    return float(value)


def workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


async def fetch_shifts(
    db: AsyncSession,
    from_date: date,
    to_date: date,
    employee_id: UUID | None = None,
    org_id: UUID | None = None,
) -> list[Shift]:
    query = (
        select(Shift)
        .options(
            selectinload(Shift.employee),
            selectinload(Shift.location),
            selectinload(Shift.work_type),
            selectinload(Shift.equipment),
        )
        .where(Shift.date >= from_date, Shift.date <= to_date)
        .order_by(Shift.date, Shift.start_time)
    )
    if org_id is not None:
        query = query.where(Shift.org_id == org_id)
    if employee_id is not None:
        query = query.where(Shift.employee_id == employee_id)
    result = await db.execute(query)
    return list(result.scalars().all())


async def build_timesheet_workbook(
    db: AsyncSession,
    from_date: date,
    to_date: date,
    employee_id: UUID | None = None,
    org_id: UUID | None = None,
) -> Workbook:
    shifts = await fetch_shifts(db, from_date, to_date, employee_id, org_id=org_id)
    workbook = new_workbook()

    detail_rows = []
    summary_map: dict[UUID, dict[str, object]] = defaultdict(
        lambda: {'shifts': 0, 'hours': 0.0, 'pay': 0.0, 'name': '', 'code': ''}
    )

    from app.services.salary import shift_pay_amount

    for shift in shifts:
        hours = shift_hours(shift)
        pay = shift_pay_amount(shift, hours)
        detail_rows.append(
            [
                fmt_date(shift.date),
                shift.employee.full_name,
                shift.employee.employee_code,
                shift.location.name,
                shift.work_type.name,
                shift.equipment.name if shift.equipment else '',
                fmt_time(shift.start_time),
                fmt_time(shift.end_time),
                hours,
                SHIFT_STATUS_LABELS.get(shift.status.value, shift.status.value),
            ]
        )
        employee_summary = summary_map[shift.employee_id]
        employee_summary['name'] = shift.employee.full_name
        employee_summary['code'] = shift.employee.employee_code
        employee_summary['shifts'] = int(employee_summary['shifts']) + 1
        employee_summary['hours'] = float(employee_summary['hours']) + hours
        employee_summary['pay'] = float(employee_summary['pay']) + pay

    ws_detail = workbook.active
    ws_detail.title = 'Табель'
    write_table(
        ws_detail,
        ['Дата', 'ФИО', 'Код', 'Объект', 'Тип работ', 'Техника', 'Начало', 'Конец', 'Часов', 'Статус'],
        detail_rows,
    )

    summary_rows = []
    total_hours = 0.0
    total_pay = 0.0
    for item in sorted(summary_map.values(), key=lambda row: str(row['name'])):
        hours = float(item['hours'])
        pay = round(float(item['pay']), 2)
        total_hours += hours
        total_pay += pay
        summary_rows.append([item['name'], item['code'], item['shifts'], hours, pay])

    ws_summary = workbook.create_sheet('Сводка')
    write_table(
        ws_summary,
        ['ФИО', 'Код', 'Смен', 'Часов', 'К выплате'],
        summary_rows,
        ['ИТОГО', '', sum(int(item['shifts']) for item in summary_map.values()), total_hours, total_pay],
    )
    return workbook


async def build_salary_workbook(
    db: AsyncSession, month: str, org_id: UUID | None = None
) -> Workbook:
    from_date, to_date = parse_month(month)
    shifts = [
        shift
        for shift in await fetch_shifts(db, from_date, to_date, org_id=org_id)
        if shift.status == ShiftStatus.closed
    ]
    employees_query = select(Employee).where(Employee.is_active.is_(True))
    if org_id is not None:
        employees_query = employees_query.where(Employee.org_id == org_id)
    employees_result = await db.execute(employees_query.order_by(Employee.full_name))
    employees = employees_result.scalars().all()

    from app.models.employee_rate import EmployeeRate
    from app.services.salary import shift_pay_amount, shift_source_label

    rates_query = (
        select(EmployeeRate)
        .options(selectinload(EmployeeRate.employee), selectinload(EmployeeRate.work_type))
        .where(
            EmployeeRate.valid_from <= to_date,
            or_(EmployeeRate.valid_to.is_(None), EmployeeRate.valid_to >= to_date),
        )
    )
    if org_id is not None:
        rates_query = rates_query.where(EmployeeRate.org_id == org_id)
    rates_result = await db.execute(
        rates_query.order_by(EmployeeRate.employee_id, EmployeeRate.valid_from.desc())
    )
    active_rates = list(rates_result.scalars().all())

    stats: dict[UUID, dict[str, object]] = {
        employee.id: {
            'name': employee.full_name,
            'code': employee.employee_code,
            'shifts': 0,
            'hours': 0.0,
            'regular_h': 0.0,
            'overtime_h': 0.0,
            'pay': 0.0,
            'notes': '',
        }
        for employee in employees
    }

    detail_rows = []
    for shift in shifts:
        hours = shift_hours(shift)
        pay = shift_pay_amount(shift, hours)
        snap = shift.rate_snapshot if isinstance(shift.rate_snapshot, dict) else {}
        regular_h = float(snap.get('regular_h', min(hours, 8.0)))
        overtime_h = float(snap.get('overtime_h', max(0.0, hours - 8.0)))
        source = shift_source_label(shift)

        detail_rows.append(
            [
                fmt_date(shift.date),
                shift.employee.full_name,
                shift.employee.employee_code,
                shift.work_type.name if shift.work_type else '',
                hours,
                regular_h,
                overtime_h,
                pay,
                source,
            ]
        )

        if shift.employee_id not in stats:
            continue
        row = stats[shift.employee_id]
        row['shifts'] = int(row['shifts']) + 1
        row['hours'] = float(row['hours']) + hours
        row['regular_h'] = float(row['regular_h']) + regular_h
        row['overtime_h'] = float(row['overtime_h']) + overtime_h
        row['pay'] = float(row['pay']) + pay

    workbook = new_workbook()
    ws_summary = workbook.active
    ws_summary.title = 'Итоги'

    summary_rows = []
    total_hours = 0.0
    total_regular = 0.0
    total_overtime = 0.0
    total_pay = 0.0
    total_shifts = 0
    for employee in employees:
        item = stats[employee.id]
        hours = round(float(item['hours']), 2)
        regular_h = round(float(item['regular_h']), 2)
        overtime_h = round(float(item['overtime_h']), 2)
        pay = round(float(item['pay']), 2)
        total_hours += hours
        total_regular += regular_h
        total_overtime += overtime_h
        total_pay += pay
        total_shifts += int(item['shifts'])
        summary_rows.append(
            [
                item['name'],
                item['code'],
                item['shifts'],
                hours,
                regular_h,
                overtime_h,
                pay,
                item['notes'],
            ]
        )

    write_table(
        ws_summary,
        [
            'Сотрудник',
            'Код',
            'Смен',
            'Часов',
            'Обычных ч',
            'Сверхурочных ч',
            'К выплате',
            'Примечание',
        ],
        summary_rows,
        ['ИТОГО', '', total_shifts, total_hours, total_regular, total_overtime, total_pay, ''],
    )

    ws_detail = workbook.create_sheet('По сменам')
    write_table(
        ws_detail,
        [
            'Дата',
            'Сотрудник',
            'Код',
            'Тип работ',
            'Часов',
            'Обычных ч',
            'Сверхурочных ч',
            'К выплате',
            'Источник',
        ],
        detail_rows,
    )

    rate_rows = [
        [
            rate.employee.full_name if rate.employee else '',
            rate.employee.employee_code if rate.employee else '',
            rate.work_type.name if rate.work_type else 'Базовая',
            to_number(rate.rate),
            to_number(rate.overtime_multiplier),
            to_number(rate.overtime_threshold_hours),
            fmt_date(rate.valid_from),
            fmt_date(rate.valid_to) if rate.valid_to else '',
            rate.notes or '',
        ]
        for rate in active_rates
    ]
    ws_rates = workbook.create_sheet('Ставки')
    write_table(
        ws_rates,
        [
            'Сотрудник',
            'Код',
            'Тип работ',
            'Ставка',
            'Множитель сверхурочных',
            'Порог ч',
            'С',
            'По',
            'Примечание',
        ],
        rate_rows,
    )
    return workbook


async def build_salary_preview(
    db: AsyncSession, month: str, org_id: UUID | None = None
) -> dict:
    from_date, to_date = parse_month(month)
    workbook_data_shifts = [
        shift
        for shift in await fetch_shifts(db, from_date, to_date, org_id=org_id)
        if shift.status == ShiftStatus.closed
    ]
    from app.services.salary import shift_pay_amount, shift_source_label

    employees_query = select(Employee).where(Employee.is_active.is_(True))
    if org_id is not None:
        employees_query = employees_query.where(Employee.org_id == org_id)
    employees_result = await db.execute(employees_query.order_by(Employee.full_name))
    employees = employees_result.scalars().all()

    rows = []
    for employee in employees:
        emp_shifts = [s for s in workbook_data_shifts if s.employee_id == employee.id]
        hours = round(sum(shift_hours(s) for s in emp_shifts), 2)
        regular_h = 0.0
        overtime_h = 0.0
        pay = 0.0
        for shift in emp_shifts:
            snap = shift.rate_snapshot if isinstance(shift.rate_snapshot, dict) else {}
            h = shift_hours(shift)
            regular_h += float(snap.get('regular_h', min(h, 8.0)))
            overtime_h += float(snap.get('overtime_h', max(0.0, h - 8.0)))
            pay += shift_pay_amount(shift, h)
        rows.append(
            {
                'employeeId': str(employee.id),
                'employeeName': employee.full_name,
                'employeeCode': employee.employee_code,
                'shiftsCount': len(emp_shifts),
                'hours': hours,
                'regularHours': round(regular_h, 2),
                'overtimeHours': round(overtime_h, 2),
                'amount': round(pay, 2),
            }
        )

    detail = [
        {
            'date': shift.date.isoformat(),
            'employeeId': str(shift.employee_id),
            'employeeName': shift.employee.full_name,
            'workType': shift.work_type.name if shift.work_type else '',
            'hours': shift_hours(shift),
            'amount': shift_pay_amount(shift),
            'source': shift_source_label(shift),
        }
        for shift in workbook_data_shifts
    ]

    return {
        'month': month,
        'from': from_date.isoformat(),
        'to': to_date.isoformat(),
        'summary': rows,
        'shifts': detail,
        'totalAmount': round(sum(item['amount'] for item in rows), 2),
    }


async def build_employee_earnings(
    db: AsyncSession, employee_id: UUID, month: str, org_id: UUID | None = None
) -> dict:
    from_date, to_date = parse_month(month)
    shifts = [
        shift
        for shift in await fetch_shifts(db, from_date, to_date, employee_id, org_id=org_id)
        if shift.status == ShiftStatus.closed
    ]
    from app.services.salary import shift_pay_amount, shift_source_label

    employee_query = select(Employee).where(Employee.id == employee_id)
    if org_id is not None:
        employee_query = employee_query.where(Employee.org_id == org_id)
    employee = (await db.execute(employee_query)).scalar_one_or_none()
    items = []
    total = 0.0
    hours_total = 0.0
    for shift in shifts:
        hours = shift_hours(shift)
        amount = shift_pay_amount(shift, hours)
        snap = shift.rate_snapshot if isinstance(shift.rate_snapshot, dict) else {}
        total += amount
        hours_total += hours
        items.append(
            {
                'date': shift.date.isoformat(),
                'workType': shift.work_type.name if shift.work_type else '',
                'hours': hours,
                'regularHours': float(snap.get('regular_h', min(hours, 8.0))),
                'overtimeHours': float(snap.get('overtime_h', max(0.0, hours - 8.0))),
                'amount': amount,
                'source': shift_source_label(shift),
            }
        )

    return {
        'month': month,
        'employeeId': str(employee_id),
        'employeeName': employee.full_name if employee else '',
        'shiftsCount': len(items),
        'hours': round(hours_total, 2),
        'totalAmount': round(total, 2),
        'shifts': items,
    }


async def build_inventory_workbook(
    db: AsyncSession, from_date: date, to_date: date, org_id: UUID | None = None
) -> Workbook:
    operations_query = (
        select(InventoryOperation)
        .join(InventoryItem, InventoryOperation.item_id == InventoryItem.id)
        .options(selectinload(InventoryOperation.item))
        .where(InventoryOperation.date >= from_date, InventoryOperation.date <= to_date)
    )
    items_query = select(InventoryItem).where(InventoryItem.is_active.is_(True))
    if org_id is not None:
        operations_query = operations_query.where(InventoryItem.org_id == org_id)
        items_query = items_query.where(InventoryItem.org_id == org_id)

    operations_result = await db.execute(
        operations_query.order_by(InventoryOperation.date.desc(), InventoryOperation.created_at.desc())
    )
    operations = operations_result.scalars().all()

    items_result = await db.execute(items_query.order_by(InventoryItem.name))
    items = items_result.scalars().all()
    inventory_labels = await load_dictionary_labels(
        db, org_id, 'inventory_category', INVENTORY_CATEGORY_LABELS
    )

    workbook = new_workbook()
    ws_ops = workbook.active
    ws_ops.title = 'Операции'
    operation_rows = [
        [
            fmt_date(operation.date),
            operation.item.name,
            OPERATION_TYPE_LABELS.get(operation.type.value, operation.type.value),
            to_number(operation.quantity),
            to_number(operation.stock_after),
            operation.reason or '',
            operation.supplier or '',
            to_number(operation.cost) if operation.cost is not None else '',
        ]
        for operation in operations
    ]
    write_table(
        ws_ops,
        ['Дата', 'Наименование', 'Тип', 'Количество', 'Остаток после', 'Причина', 'Поставщик', 'Стоимость'],
        operation_rows,
    )

    ws_stock = workbook.create_sheet('Остатки')
    stock_rows = [
        [
            item.name,
            resolve_label(
                str(getattr(item.category, 'value', item.category)),
                inventory_labels,
            ),
            item.unit,
            to_number(item.current_stock),
            to_number(item.min_stock),
            'Да' if item.current_stock < item.min_stock else 'Нет',
        ]
        for item in items
    ]
    write_table(
        ws_stock,
        ['Наименование', 'Категория', 'Ед.', 'Остаток', 'Мин.', 'Критично'],
        stock_rows,
    )
    return workbook


async def build_shipments_workbook(
    db: AsyncSession, from_date: date, to_date: date, org_id: UUID | None = None
) -> Workbook:
    query = select(Shipment).where(Shipment.date >= from_date, Shipment.date <= to_date)
    if org_id is not None:
        query = query.where(Shipment.org_id == org_id)
    result = await db.execute(query.order_by(Shipment.date))
    shipments = result.scalars().all()

    workbook = new_workbook()
    ws = workbook.active
    ws.title = 'Отгрузки'

    rows = []
    total_kg = 0.0
    total_sum = 0.0
    for shipment in shipments:
        kg = to_number(shipment.quantity_kg)
        price = to_number(shipment.price_per_kg) if shipment.price_per_kg is not None else None
        amount = round(kg * price, 2) if price is not None else 0.0
        total_kg += kg
        total_sum += amount
        rows.append(
            [
                fmt_date(shipment.date),
                shipment.crop_type,
                kg,
                shipment.destination or '',
                price if price is not None else '',
                amount if price is not None else '',
            ]
        )

    write_table(
        ws,
        ['Дата', 'Культура', 'Кг', 'Направление', 'Цена/кг', 'Сумма'],
        rows,
        ['ИТОГО', '', total_kg, '', '', total_sum],
    )
    return workbook


async def build_expenses_workbook(
    db: AsyncSession, from_date: date, to_date: date, org_id: UUID | None = None
) -> Workbook:
    query = select(Expense).where(Expense.date >= from_date, Expense.date <= to_date)
    if org_id is not None:
        query = query.where(Expense.org_id == org_id)
    result = await db.execute(query.order_by(Expense.date))
    expenses = result.scalars().all()
    category_labels = await load_dictionary_labels(
        db, org_id, 'expense_category', EXPENSE_CATEGORY_LABELS
    )

    workbook = new_workbook()
    ws = workbook.active
    ws.title = 'Затраты'

    rows = []
    total_amount = 0.0
    for expense in expenses:
        amount = to_number(expense.amount)
        total_amount += amount
        rows.append(
            [
                fmt_date(expense.date),
                resolve_label(expense.category, category_labels),
                amount,
                expense.description or '',
                expense.supplier or '',
            ]
        )

    write_table(
        ws,
        ['Дата', 'Категория', 'Сумма', 'Описание', 'Поставщик'],
        rows,
        ['ИТОГО', '', total_amount, '', ''],
    )
    return workbook


async def build_summary_workbook(
    db: AsyncSession, month: str, org_id: UUID | None = None
) -> Workbook:
    from_date, to_date = parse_month(month)
    shifts = await fetch_shifts(db, from_date, to_date, org_id=org_id)

    shipments_query = select(Shipment).where(
        Shipment.date >= from_date, Shipment.date <= to_date
    )
    if org_id is not None:
        shipments_query = shipments_query.where(Shipment.org_id == org_id)
    shipments_result = await db.execute(shipments_query)
    shipments = shipments_result.scalars().all()

    expenses_query = select(Expense).where(
        Expense.date >= from_date, Expense.date <= to_date
    )
    if org_id is not None:
        expenses_query = expenses_query.where(Expense.org_id == org_id)
    expenses_result = await db.execute(expenses_query)
    expenses = expenses_result.scalars().all()

    operations_query = (
        select(InventoryOperation)
        .join(InventoryItem, InventoryOperation.item_id == InventoryItem.id)
        .where(
            InventoryOperation.date >= from_date,
            InventoryOperation.date <= to_date,
        )
    )
    if org_id is not None:
        operations_query = operations_query.where(InventoryItem.org_id == org_id)
    operations_result = await db.execute(operations_query)
    operations = operations_result.scalars().all()

    items_query = select(InventoryItem).where(InventoryItem.is_active.is_(True))
    if org_id is not None:
        items_query = items_query.where(InventoryItem.org_id == org_id)
    items_result = await db.execute(items_query)
    items = items_result.scalars().all()

    category_labels = await load_dictionary_labels(
        db, org_id, 'expense_category', EXPENSE_CATEGORY_LABELS
    )

    workbook = new_workbook()

    total_hours = sum(shift_hours(shift) for shift in shifts)
    employee_hours: dict[str, float] = defaultdict(float)
    for shift in shifts:
        employee_hours[shift.employee.full_name] += shift_hours(shift)

    ws_time = workbook.active
    ws_time.title = 'Рабочее время'
    time_rows = [
        ['Период', f'{fmt_date(from_date)} - {fmt_date(to_date)}'],
        ['Всего смен', len(shifts)],
        ['Всего часов', round(total_hours, 2)],
        ['Сотрудников', len(employee_hours)],
    ]
    write_table(ws_time, ['Показатель', 'Значение'], time_rows)
    employee_rows = [
        [name, round(hours, 2)] for name, hours in sorted(employee_hours.items())
    ]
    start_row = ws_time.max_row + 2
    ws_time.cell(row=start_row, column=1, value='ФИО')
    ws_time.cell(row=start_row, column=2, value='Часов')
    for offset, row in enumerate(employee_rows, start=1):
        ws_time.cell(row=start_row + offset, column=1, value=row[0])
        ws_time.cell(row=start_row + offset, column=2, value=row[1])

    shipments_kg = sum(to_number(item.quantity_kg) for item in shipments)
    shipments_sum = sum(
        to_number(item.quantity_kg) * to_number(item.price_per_kg)
        for item in shipments
        if item.price_per_kg is not None
    )
    expenses_sum = sum(to_number(item.amount) for item in expenses)
    expenses_by_category: dict[str, float] = defaultdict(float)
    for expense in expenses:
        label = resolve_label(expense.category, category_labels)
        expenses_by_category[label] += to_number(expense.amount)

    ws_finance = workbook.create_sheet('Финансы')
    finance_rows = [
        ['Отгрузки (кг)', shipments_kg],
        ['Отгрузки (сумма)', round(shipments_sum, 2)],
        ['Затраты (всего)', round(expenses_sum, 2)],
        ['Баланс', round(shipments_sum - expenses_sum, 2)],
    ]
    write_table(ws_finance, ['Показатель', 'Сумма'], finance_rows)
    category_start = ws_finance.max_row + 2
    ws_finance.cell(row=category_start, column=1, value='Категория')
    ws_finance.cell(row=category_start, column=2, value='Сумма')
    for offset, (category, amount) in enumerate(sorted(expenses_by_category.items()), start=1):
        ws_finance.cell(row=category_start + offset, column=1, value=category)
        ws_finance.cell(row=category_start + offset, column=2, value=round(amount, 2))

    income_count = sum(1 for op in operations if op.type.value == 'income')
    expense_count = sum(1 for op in operations if op.type.value == 'expense')
    critical_count = sum(1 for item in items if item.current_stock < item.min_stock)

    ws_stock = workbook.create_sheet('Склад')
    stock_rows = [
        ['Операций приход', income_count],
        ['Операций расход', expense_count],
        ['Критичных позиций', critical_count],
        ['Всего позиций', len(items)],
    ]
    write_table(ws_stock, ['Показатель', 'Значение'], stock_rows)
    critical_rows = [
        [item.name, to_number(item.current_stock), to_number(item.min_stock), item.unit]
        for item in items
        if item.current_stock < item.min_stock
    ]
    critical_start = ws_stock.max_row + 2
    ws_stock.cell(row=critical_start, column=1, value='Наименование')
    ws_stock.cell(row=critical_start, column=2, value='Остаток')
    ws_stock.cell(row=critical_start, column=3, value='Мин.')
    ws_stock.cell(row=critical_start, column=4, value='Ед.')
    for offset, row in enumerate(critical_rows, start=1):
        for col, value in enumerate(row, start=1):
            ws_stock.cell(row=critical_start + offset, column=col, value=value)

    return workbook


async def fetch_equipment_list(
    db: AsyncSession,
    equipment_id: UUID | None = None,
    org_id: UUID | None = None,
) -> list[Equipment]:
    query = select(Equipment).where(Equipment.is_active.is_(True)).order_by(Equipment.name)
    if org_id is not None:
        query = query.where(Equipment.org_id == org_id)
    if equipment_id is not None:
        query = query.where(Equipment.id == equipment_id)
    result = await db.execute(query)
    return list(result.scalars().all())


async def build_equipment_workbook(
    db: AsyncSession,
    from_date: date,
    to_date: date,
    equipment_id: UUID | None = None,
    org_id: UUID | None = None,
) -> Workbook:
    equipment_list = await fetch_equipment_list(db, equipment_id, org_id=org_id)
    equipment_ids = [item.id for item in equipment_list]
    category_labels = await load_dictionary_labels(
        db, org_id, 'expense_category', EXPENSE_CATEGORY_LABELS
    )

    meter_logs: list[EquipmentMeterLog] = []
    if equipment_ids:
        logs_result = await db.execute(
            select(EquipmentMeterLog)
            .options(
                selectinload(EquipmentMeterLog.equipment),
                selectinload(EquipmentMeterLog.created_by_user),
                selectinload(EquipmentMeterLog.shift).selectinload(Shift.employee),
            )
            .where(
                EquipmentMeterLog.date >= from_date,
                EquipmentMeterLog.date <= to_date,
                EquipmentMeterLog.equipment_id.in_(equipment_ids),
            )
            .order_by(EquipmentMeterLog.date, EquipmentMeterLog.created_at)
        )
        meter_logs = list(logs_result.scalars().all())

    maintenance_records: list[EquipmentMaintenance] = []
    if equipment_ids:
        maintenance_result = await db.execute(
            select(EquipmentMaintenance)
            .options(selectinload(EquipmentMaintenance.equipment))
            .where(
                EquipmentMaintenance.date >= from_date,
                EquipmentMaintenance.date <= to_date,
                EquipmentMaintenance.equipment_id.in_(equipment_ids),
            )
            .order_by(EquipmentMaintenance.date)
        )
        maintenance_records = list(maintenance_result.scalars().all())

    expenses: list[Expense] = []
    if equipment_ids:
        expenses_result = await db.execute(
            select(Expense)
            .options(selectinload(Expense.equipment))
            .where(
                Expense.date >= from_date,
                Expense.date <= to_date,
                Expense.equipment_id.in_(equipment_ids),
            )
            .order_by(Expense.date)
        )
        expenses = list(expenses_result.scalars().all())

    logs_by_equipment: dict[UUID, list[EquipmentMeterLog]] = defaultdict(list)
    for log in meter_logs:
        logs_by_equipment[log.equipment_id].append(log)

    maintenance_by_equipment: dict[UUID, list[EquipmentMaintenance]] = defaultdict(list)
    for record in maintenance_records:
        maintenance_by_equipment[record.equipment_id].append(record)

    expenses_by_equipment: dict[UUID, list[Expense]] = defaultdict(list)
    for expense in expenses:
        if expense.equipment_id is not None:
            expenses_by_equipment[expense.equipment_id].append(expense)

    workbook = new_workbook()

    summary_rows = []
    total_added = 0.0
    total_to = 0
    total_expenses = 0.0
    for equipment in equipment_list:
        added = sum(to_number(log.value_added) for log in logs_by_equipment.get(equipment.id, []))
        end_meter = to_number(equipment.current_meter)
        start_meter = round(end_meter - added, 2)
        to_count = len(maintenance_by_equipment.get(equipment.id, []))
        expense_sum = sum(
            to_number(item.amount) for item in expenses_by_equipment.get(equipment.id, [])
        )
        total_added += added
        total_to += to_count
        total_expenses += expense_sum
        summary_rows.append(
            [
                equipment.name,
                calc_meter_label(equipment.meter_type),
                start_meter,
                end_meter,
                round(added, 2),
                to_count,
                round(expense_sum, 2),
            ]
        )

    ws_summary = workbook.active
    ws_summary.title = 'Сводка'
    write_table(
        ws_summary,
        [
            'Техника',
            'Тип счётчика',
            'Показатель нач.',
            'Показатель кон.',
            'Добавлено',
            'ТО за период',
            'Затраты ₽',
        ],
        summary_rows,
        ['ИТОГО', '', '', '', round(total_added, 2), total_to, round(total_expenses, 2)],
    )

    def shift_label(log: EquipmentMeterLog) -> str:
        if log.shift is None:
            return ''
        employee = log.shift.employee.full_name if log.shift.employee else ''
        return f'{fmt_date(log.shift.date)} {employee}'.strip()

    ws_logs = workbook.create_sheet('Журнал показаний')
    log_rows = [
        [
            fmt_date(log.date),
            log.equipment.name if log.equipment else '',
            to_number(log.value_added),
            to_number(log.meter_after),
            shift_label(log),
            log.note or '',
            log.created_by_user.full_name if log.created_by_user else '',
        ]
        for log in meter_logs
    ]
    write_table(
        ws_logs,
        ['Дата', 'Техника', 'Добавлено', 'Итого', 'Смена', 'Примечание', 'Внёс'],
        log_rows,
    )

    ws_maintenance = workbook.create_sheet('История ТО')
    maintenance_rows = []
    for record in maintenance_records:
        next_interval = ''
        if record.next_to_at is not None and record.meter_at is not None:
            next_interval = round(
                to_number(record.next_to_at) - to_number(record.meter_at),
                2,
            )
        maintenance_rows.append(
            [
                fmt_date(record.date),
                record.equipment.name if record.equipment else '',
                record.type,
                to_number(record.meter_at) if record.meter_at is not None else '',
                to_number(record.cost) if record.cost is not None else '',
                next_interval,
                record.description or '',
            ]
        )
    write_table(
        ws_maintenance,
        [
            'Дата',
            'Техника',
            'Тип ТО',
            'При показателе',
            'Стоимость ₽',
            'До следующего',
            'Описание',
        ],
        maintenance_rows,
    )

    ws_expenses = workbook.create_sheet('Затраты на технику')
    expense_rows = [
        [
            fmt_date(expense.date),
            expense.equipment.name if expense.equipment else '',
            resolve_label(expense.category, category_labels),
            to_number(expense.amount),
            expense.description or '',
            expense.supplier or '',
        ]
        for expense in expenses
    ]
    write_table(
        ws_expenses,
        ['Дата', 'Техника', 'Категория', 'Сумма ₽', 'Описание', 'Поставщик'],
        expense_rows,
        ['ИТОГО', '', '', round(total_expenses, 2), '', ''],
    )

    implements_query = (
        select(Implement)
        .options(selectinload(Implement.current_equipment))
        .where(Implement.is_active.is_(True))
        .order_by(Implement.name)
    )
    if equipment_id is not None:
        implements_query = implements_query.where(Implement.current_equipment_id == equipment_id)
    implements_result = await db.execute(implements_query)
    implements = list(implements_result.scalars().all())
    implement_ids = [item.id for item in implements]

    implement_maintenance: list[ImplementMaintenance] = []
    if implement_ids:
        impl_maint_result = await db.execute(
            select(ImplementMaintenance).where(
                ImplementMaintenance.implement_id.in_(implement_ids),
                ImplementMaintenance.date >= from_date,
                ImplementMaintenance.date <= to_date,
            )
        )
        implement_maintenance = list(impl_maint_result.scalars().all())

    maint_by_implement: dict[UUID, list[ImplementMaintenance]] = defaultdict(list)
    for record in implement_maintenance:
        maint_by_implement[record.implement_id].append(record)

    ws_implements = workbook.create_sheet('Приспособления')
    implement_rows = []
    for implement in implements:
        maint_records = maint_by_implement.get(implement.id, [])
        maint_cost = sum(to_number(item.cost) for item in maint_records if item.cost is not None)
        implement_rows.append(
            [
                implement.name,
                implement.category,
                IMPLEMENT_CONDITION_LABELS.get(implement.condition, implement.condition),
                implement.current_equipment.name if implement.current_equipment else '—',
                len(maint_records),
                round(maint_cost, 2),
            ]
        )
    write_table(
        ws_implements,
        [
            'Название',
            'Категория',
            'Состояние',
            'Привязана к технике',
            'ТО за период',
            'Затраты ₽',
        ],
        implement_rows,
    )

    return workbook


async def fetch_field_locations(
    db: AsyncSession,
    field_id: UUID | None = None,
    org_id: UUID | None = None,
) -> list[Location]:
    query = (
        select(Location)
        .where(
            Location.is_active.is_(True),
            or_(Location.crop_type.is_not(None), Location.name.like('Поле%')),
        )
        .order_by(Location.name)
    )
    if org_id is not None:
        query = query.where(Location.org_id == org_id)
    if field_id is not None:
        query = query.where(Location.id == field_id)
    result = await db.execute(query)
    return list(result.scalars().all())


async def fetch_field_shifts(
    db: AsyncSession,
    from_date: date,
    to_date: date,
    field_id: UUID | None = None,
    org_id: UUID | None = None,
) -> list[Shift]:
    query = (
        select(Shift)
        .options(
            selectinload(Shift.employee),
            selectinload(Shift.field),
            selectinload(Shift.work_type),
            selectinload(Shift.equipment),
            selectinload(Shift.implement),
        )
        .where(
            Shift.date >= from_date,
            Shift.date <= to_date,
            Shift.field_id.is_not(None),
        )
        .order_by(Shift.date, Shift.start_time)
    )
    if org_id is not None:
        query = query.where(Shift.org_id == org_id)
    if field_id is not None:
        query = query.where(Shift.field_id == field_id)
    result = await db.execute(query)
    return list(result.scalars().all())


def unique_join(values: list[str]) -> str:
    seen: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
    return ', '.join(seen)


async def build_fields_workbook(
    db: AsyncSession,
    from_date: date,
    to_date: date,
    field_id: UUID | None = None,
    org_id: UUID | None = None,
) -> Workbook:
    fields = await fetch_field_locations(db, field_id, org_id=org_id)
    field_map = {item.id: item for item in fields}
    shifts = await fetch_field_shifts(db, from_date, to_date, field_id, org_id=org_id)

    shifts_by_field: dict[UUID, list[Shift]] = defaultdict(list)
    for shift in shifts:
        if shift.field_id is not None:
            shifts_by_field[shift.field_id].append(shift)

    if field_id is not None:
        field_ids = list(field_map.keys())
    else:
        field_ids = list(dict.fromkeys(list(field_map.keys()) + list(shifts_by_field.keys())))

    workbook = new_workbook()
    summary_rows = []
    for fid in field_ids:
        field = field_map.get(fid)
        if field is None and shifts_by_field.get(fid):
            field = shifts_by_field[fid][0].field
        if field is None:
            continue
        field_shifts = shifts_by_field.get(fid, [])
        equipment_names = unique_join(
            [shift.equipment.name for shift in field_shifts if shift.equipment]
        )
        employee_names = unique_join(
            [shift.employee.full_name for shift in field_shifts if shift.employee]
        )
        total_hours = round(sum(shift_hours(shift) for shift in field_shifts), 2)
        summary_rows.append(
            [
                field.name,
                field.crop_type or '',
                to_number(field.area_ha) if field.area_ha is not None else '',
                len(field_shifts),
                total_hours,
                equipment_names,
                employee_names,
            ]
        )

    ws_summary = workbook.active
    ws_summary.title = 'Работы по полям'
    write_table(
        ws_summary,
        [
            'Поле',
            'Культура',
            'Площадь га',
            'Кол-во смен',
            'Часов отработано',
            'Техника (список)',
            'Сотрудники (список)',
        ],
        summary_rows,
    )

    ws_detail = workbook.create_sheet('Детальный журнал смен')
    detail_rows = [
        [
            fmt_date(shift.date),
            shift.field.name if shift.field else '',
            shift.employee.full_name if shift.employee else '',
            shift.work_type.name if shift.work_type else '',
            shift.equipment.name if shift.equipment else '',
            shift.implement.name if shift.implement else '',
            shift_hours(shift),
        ]
        for shift in shifts
    ]
    write_table(
        ws_detail,
        ['Дата', 'Поле', 'Сотрудник', 'Тип работы', 'Техника', 'Приспособление', 'Часов'],
        detail_rows,
    )

    agro_query = (
        select(AgroPlan)
        .options(
            selectinload(AgroPlan.location),
            selectinload(AgroPlan.work_type),
            selectinload(AgroPlan.equipment),
            selectinload(AgroPlan.actual_shift),
        )
        .where(
            AgroPlan.planned_date >= from_date,
            AgroPlan.planned_date <= to_date,
        )
        .order_by(AgroPlan.planned_date)
    )
    if field_id is not None:
        agro_query = agro_query.where(AgroPlan.location_id == field_id)
    agro_result = await db.execute(agro_query)
    plans = list(agro_result.scalars().all())

    ws_agro = workbook.create_sheet('Агрокалендарь')
    agro_rows = []
    for plan in plans:
        actual_date = ''
        if plan.actual_shift is not None:
            actual_date = fmt_date(plan.actual_shift.date)
        agro_rows.append(
            [
                plan.location.name if plan.location else '',
                plan.work_type.name if plan.work_type else '',
                fmt_date(plan.planned_date),
                actual_date,
                AGRO_STATUS_LABELS.get(plan.status, plan.status),
                plan.equipment.name if plan.equipment else '',
                plan.notes or '',
            ]
        )
    write_table(
        ws_agro,
        ['Поле', 'Тип работы', 'Дата план', 'Дата факт', 'Статус', 'Техника', 'Примечания'],
        agro_rows,
    )

    return workbook


def year_range(year: int) -> tuple[date, date]:
    return date(year, 1, 1), date(year, 12, 31)


async def build_season_workbook(
    db: AsyncSession, year: int, org_id: UUID | None = None
) -> Workbook:
    from_date, to_date = year_range(year)
    fields = await fetch_field_locations(db, org_id=org_id)
    shifts = await fetch_field_shifts(db, from_date, to_date, org_id=org_id)

    hours_by_field_month: dict[UUID, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for shift in shifts:
        if shift.field_id is None:
            continue
        hours_by_field_month[shift.field_id][shift.date.month] += shift_hours(shift)

    workbook = new_workbook()
    matrix_headers = ['Поле', 'Культура', 'Площадь га', *MONTH_LABELS, 'Итого']
    matrix_rows = []
    for field in fields:
        month_hours = hours_by_field_month.get(field.id, {})
        row_total = 0.0
        month_values = []
        for month in range(1, 13):
            value = round(month_hours.get(month, 0.0), 2)
            month_values.append(value)
            row_total += value
        matrix_rows.append(
            [
                field.name,
                field.crop_type or '',
                to_number(field.area_ha) if field.area_ha is not None else '',
                *month_values,
                round(row_total, 2),
            ]
        )

    ws_matrix = workbook.active
    ws_matrix.title = 'Работы по месяцам'
    write_table(ws_matrix, matrix_headers, matrix_rows)

    equipment_data = await build_equipment_workbook(db, from_date, to_date, None, org_id=org_id)
    ws_equipment_src = equipment_data['Сводка']
    equipment_rows = []
    for row in ws_equipment_src.iter_rows(
        min_row=2,
        max_row=ws_equipment_src.max_row - 1,
        values_only=True,
    ):
        if row and row[0] != 'ИТОГО':
            equipment_rows.append([row[0], row[2], row[3], row[4], row[5], row[6]])

    ws_equipment = workbook.create_sheet('Техника за год')
    write_table(
        ws_equipment,
        [
            'Техника',
            'Показатель нач. года',
            'Показатель кон.',
            'Добавлено',
            'ТО за год',
            'Затраты ₽',
        ],
        equipment_rows,
    )

    expenses_query = select(Expense).where(Expense.date >= from_date, Expense.date <= to_date)
    if org_id is not None:
        expenses_query = expenses_query.where(Expense.org_id == org_id)
    expenses_result = await db.execute(expenses_query)
    expenses = list(expenses_result.scalars().all())

    shipments_query = select(Shipment).where(Shipment.date >= from_date, Shipment.date <= to_date)
    if org_id is not None:
        shipments_query = shipments_query.where(Shipment.org_id == org_id)
    shipments_result = await db.execute(shipments_query)
    shipments = list(shipments_result.scalars().all())

    finance_rows = []
    total_expenses = 0.0
    total_revenue = 0.0
    total_net = 0.0
    for month in range(1, 13):
        month_expenses = sum(
            to_number(item.amount) for item in expenses if item.date.month == month
        )
        month_revenue = sum(
            to_number(item.quantity_kg) * to_number(item.price_per_kg)
            for item in shipments
            if item.date.month == month and item.price_per_kg is not None
        )
        month_net = round(month_revenue - month_expenses, 2)
        total_expenses += month_expenses
        total_revenue += month_revenue
        total_net += month_net
        finance_rows.append(
            [
                MONTH_LABELS[month - 1],
                round(month_expenses, 2),
                round(month_revenue, 2),
                month_net,
            ]
        )

    ws_finance = workbook.create_sheet('Финансы по месяцам')
    write_table(
        ws_finance,
        ['Месяц', 'Затраты', 'Отгрузки (выручка)', 'Итого'],
        finance_rows,
        ['ИТОГО', round(total_expenses, 2), round(total_revenue, 2), round(total_net, 2)],
    )

    year_shifts = await fetch_shifts(db, from_date, to_date, org_id=org_id)
    employees_query = select(Employee).where(Employee.is_active.is_(True))
    if org_id is not None:
        employees_query = employees_query.where(Employee.org_id == org_id)
    employees_result = await db.execute(employees_query.order_by(Employee.full_name))
    employees = list(employees_result.scalars().all())

    from app.services.salary import shift_pay_amount

    employee_stats: dict[UUID, dict[str, object]] = {
        employee.id: {
            'shifts': 0,
            'hours': 0.0,
            'pay': 0.0,
            'name': employee.full_name,
        }
        for employee in employees
    }
    for shift in year_shifts:
        if shift.employee_id not in employee_stats:
            continue
        hours = shift_hours(shift)
        employee_stats[shift.employee_id]['shifts'] = (
            int(employee_stats[shift.employee_id]['shifts']) + 1
        )
        employee_stats[shift.employee_id]['hours'] = (
            float(employee_stats[shift.employee_id]['hours']) + hours
        )
        employee_stats[shift.employee_id]['pay'] = (
            float(employee_stats[shift.employee_id]['pay']) + shift_pay_amount(shift, hours)
        )

    employee_rows = []
    total_hours = 0.0
    total_pay = 0.0
    for employee in employees:
        stats = employee_stats[employee.id]
        hours = round(float(stats['hours']), 2)
        pay = round(float(stats['pay']), 2)
        total_hours += hours
        total_pay += pay
        employee_rows.append([stats['name'], stats['shifts'], hours, pay])

    ws_employees = workbook.create_sheet('Сотрудники за год')
    write_table(
        ws_employees,
        ['ФИО', 'Смен за год', 'Часов за год', 'К выплате ₽'],
        employee_rows,
        ['ИТОГО', sum(int(employee_stats[e.id]['shifts']) for e in employees), total_hours, total_pay],
    )

    return workbook
