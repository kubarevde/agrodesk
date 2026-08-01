"""Holding reports overlay — reuse build_*_workbook; do not widen /api/reports."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from typing import Any, Awaitable, Callable, Literal
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl.workbook.workbook import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.organization import Organization
from app.services.excel_styles import new_workbook, write_table
from app.services.holding import list_holding_children, require_head_org
from app.services.reports import (
    build_equipment_workbook,
    build_expenses_workbook,
    build_fields_workbook,
    build_inventory_workbook,
    build_maintenance_workbook,
    build_purchases_workbook,
    build_salary_workbook,
    build_season_workbook,
    build_shipment_requests_workbook,
    build_shipments_workbook,
    build_summary_workbook,
    build_timesheet_workbook,
)

HoldingReportMode = Literal['child', 'group']
PeriodMode = Literal['range', 'month', 'year']

BuilderFn = Callable[..., Awaitable[Workbook]]


@dataclass(frozen=True)
class HoldingReportSpec:
    report_id: str
    title: str
    period_mode: PeriodMode
    modes: tuple[HoldingReportMode, ...]
    reason_if_no_group: str | None = None


# Whitelist: group = honest sums/counts or multi-sheet per child (no PII merge).
# child-only = payroll / people / local assets — export one child via builders.
HOLDING_REPORT_SPECS: dict[str, HoldingReportSpec] = {
    'shipments': HoldingReportSpec(
        'shipments', 'Отгрузки урожая', 'range', ('child', 'group')
    ),
    'expenses': HoldingReportSpec(
        'expenses', 'Затраты', 'range', ('child', 'group')
    ),
    'summary': HoldingReportSpec(
        'summary', 'Сводный KPI', 'month', ('child', 'group')
    ),
    'inventory': HoldingReportSpec(
        'inventory',
        'Склад ТМЦ',
        'range',
        ('child', 'group'),
    ),
    'purchases': HoldingReportSpec(
        'purchases', 'Закупки', 'range', ('child', 'group')
    ),
    'maintenance': HoldingReportSpec(
        'maintenance', 'Ремонт и обслуживание', 'range', ('child', 'group')
    ),
    'timesheet': HoldingReportSpec(
        'timesheet',
        'Табель',
        'range',
        ('child',),
        reason_if_no_group='Содержит персональные смены и ФИО — только одна КФХ',
    ),
    'salary': HoldingReportSpec(
        'salary',
        'Зарплатная ведомость',
        'month',
        ('child',),
        reason_if_no_group='Расчёт зарплаты и ставки — только одна КФХ',
    ),
    'shipment-requests': HoldingReportSpec(
        'shipment-requests',
        'Заявки на отгрузку',
        'range',
        ('child',),
        reason_if_no_group='Исполнители и заказчики — только одна КФХ',
    ),
    'equipment': HoldingReportSpec(
        'equipment',
        'Техника и ресурс',
        'range',
        ('child',),
        reason_if_no_group='Парк и наработка не суммируются между КФХ',
    ),
    'fields': HoldingReportSpec(
        'fields',
        'Отчёт по полям',
        'range',
        ('child',),
        reason_if_no_group='Поля и журнал смен локальны для КФХ',
    ),
    'season': HoldingReportSpec(
        'season',
        'Сезонный обзор',
        'year',
        ('child',),
        reason_if_no_group='Включает зарплату сотрудников и локальный парк',
    ),
}


def catalog_payload() -> list[dict[str, Any]]:
    return [
        {
            'report_id': spec.report_id,
            'title': spec.title,
            'period_mode': spec.period_mode,
            'modes': list(spec.modes),
            'group_unsupported_reason': spec.reason_if_no_group,
        }
        for spec in HOLDING_REPORT_SPECS.values()
    ]


def _safe_sheet_title(prefix: str, sheet_title: str, used: set[str]) -> str:
    # Excel forbids: \ / ? * [ ] :
    base = f'{prefix}-{sheet_title}'.replace(':', '-')[:31]
    title = base
    n = 2
    while title in used:
        suffix = f'~{n}'
        title = f'{base[: 31 - len(suffix)]}{suffix}'
        n += 1
    used.add(title)
    return title


def _copy_sheet_values(source, target) -> None:
    for row in source.iter_rows():
        for cell in row:
            target.cell(row=cell.row, column=cell.column, value=cell.value)


def prepend_scope_sheet(workbook: Workbook, rows: list[list[object]]) -> Workbook:
    ws = workbook.create_sheet('Область', 0)
    write_table(ws, ['Параметр', 'Значение'], rows)
    return workbook


def merge_org_workbooks(
    parts: list[tuple[str, str, Workbook]],
    *,
    report_title: str,
) -> Workbook:
    """Combine per-child workbooks; each sheet labeled with org slug (Excel 31-char titles)."""
    master = new_workbook()
    cover = master.active
    cover.title = 'Область'
    write_table(
        cover,
        ['Параметр', 'Значение'],
        [
            ['Тип', 'Сводка по КФХ (holding)'],
            ['Отчёт', report_title],
            ['КФХ', ', '.join(name for name, _, _ in parts)],
            ['Примечание', 'Листы помечены slug организации; marketplace не включён'],
        ],
    )
    used: set[str] = {'Область'}
    for name, slug, wb in parts:
        prefix = (slug or name)[:12]
        for sheet in wb.worksheets:
            title = _safe_sheet_title(prefix, sheet.title, used)
            dest = master.create_sheet(title)
            _copy_sheet_values(sheet, dest)
    return master


async def _resolve_targets(
    db: AsyncSession,
    *,
    head_org_id: UUID,
    mode: HoldingReportMode,
    child_org_id: UUID | None,
) -> list[Organization]:
    await require_head_org(db, head_org_id)
    views = await list_holding_children(db, head_org_id)
    by_id = {v.org_id: v for v in views}

    if mode == 'child':
        if child_org_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Укажите child_org_id',
            )
        if child_org_id not in by_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail='Организация не является дочерней для текущей головной',
            )
        org = await db.get(Organization, child_org_id)
        if org is None or not org.is_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Дочерняя организация недоступна',
            )
        return [org]

    # group
    orgs: list[Organization] = []
    for view in views:
        org = await db.get(Organization, view.org_id)
        if org is not None and org.is_active:
            orgs.append(org)
    if not orgs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Нет активных дочерних КФХ',
        )
    return orgs


async def _build_for_org(
    db: AsyncSession,
    *,
    report_id: str,
    org_id: UUID,
    from_date: date | None,
    to_date: date | None,
    month: str | None,
    year: int | None,
) -> Workbook:
    if report_id == 'shipments':
        assert from_date and to_date
        return await build_shipments_workbook(db, from_date, to_date, org_id=org_id)
    if report_id == 'expenses':
        assert from_date and to_date
        return await build_expenses_workbook(db, from_date, to_date, org_id=org_id)
    if report_id == 'summary':
        assert month
        return await build_summary_workbook(db, month, org_id=org_id)
    if report_id == 'inventory':
        assert from_date and to_date
        return await build_inventory_workbook(db, from_date, to_date, org_id=org_id)
    if report_id == 'purchases':
        assert from_date and to_date
        return await build_purchases_workbook(db, from_date, to_date, org_id=org_id)
    if report_id == 'maintenance':
        assert from_date and to_date
        return await build_maintenance_workbook(db, from_date, to_date, org_id=org_id)
    if report_id == 'timesheet':
        assert from_date and to_date
        return await build_timesheet_workbook(
            db, from_date, to_date, employee_id=None, org_id=org_id
        )
    if report_id == 'salary':
        assert month
        return await build_salary_workbook(db, month, org_id=org_id)
    if report_id == 'shipment-requests':
        assert from_date and to_date
        return await build_shipment_requests_workbook(
            db, from_date, to_date, org_id=org_id
        )
    if report_id == 'equipment':
        assert from_date and to_date
        return await build_equipment_workbook(
            db, from_date, to_date, equipment_id=None, org_id=org_id
        )
    if report_id == 'fields':
        assert from_date and to_date
        return await build_fields_workbook(
            db, from_date, to_date, field_id=None, org_id=org_id
        )
    if report_id == 'season':
        assert year is not None
        return await build_season_workbook(db, year, org_id=org_id)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail='Неизвестный отчёт',
    )


def _validate_period(spec: HoldingReportSpec, **kwargs: Any) -> None:
    if spec.period_mode == 'range':
        if not kwargs.get('from_date') or not kwargs.get('to_date'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Нужны from_date и to_date',
            )
    elif spec.period_mode == 'month':
        if not kwargs.get('month'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Нужен month (YYYY-MM)',
            )
    elif spec.period_mode == 'year':
        if kwargs.get('year') is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Нужен year',
            )


async def build_holding_report_workbook(
    db: AsyncSession,
    *,
    head_org_id: UUID,
    report_id: str,
    mode: HoldingReportMode,
    child_org_id: UUID | None,
    from_date: date | None = None,
    to_date: date | None = None,
    month: str | None = None,
    year: int | None = None,
) -> tuple[Workbook, str]:
    spec = HOLDING_REPORT_SPECS.get(report_id)
    if spec is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Отчёт не поддерживается в holding mode',
        )
    if mode not in spec.modes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=spec.reason_if_no_group
            or 'Этот отчёт нельзя агрегировать по группе КФХ',
        )
    _validate_period(
        spec, from_date=from_date, to_date=to_date, month=month, year=year
    )

    targets = await _resolve_targets(
        db, head_org_id=head_org_id, mode=mode, child_org_id=child_org_id
    )

    if mode == 'child':
        org = targets[0]
        wb = await _build_for_org(
            db,
            report_id=report_id,
            org_id=org.id,
            from_date=from_date,
            to_date=to_date,
            month=month,
            year=year,
        )
        prepend_scope_sheet(
            wb,
            [
                ['Тип', 'Отчёт по одной КФХ (holding)'],
                ['Отчёт', spec.title],
                ['КФХ', org.name],
                ['Slug', org.slug],
                ['org_id', str(org.id)],
            ],
        )
        filename = f'holding_{report_id}_{org.slug}.xlsx'
        return wb, filename

    parts: list[tuple[str, str, Workbook]] = []
    for org in targets:
        wb = await _build_for_org(
            db,
            report_id=report_id,
            org_id=org.id,
            from_date=from_date,
            to_date=to_date,
            month=month,
            year=year,
        )
        parts.append((org.name, org.slug, wb))
    merged = merge_org_workbooks(parts, report_title=spec.title)
    return merged, f'holding_{report_id}_group.xlsx'
