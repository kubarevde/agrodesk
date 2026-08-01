"""Holding reports overlay — reuse builders; do not widen /api/reports."""

from __future__ import annotations

import ast
import asyncio
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import httpx
import openpyxl
import pytest
from sqlalchemy import delete, or_
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import settings
from app.models.org_hierarchy import OrgHierarchyLink
from app.models.organization import Organization
from app.services.holding_reports import HOLDING_REPORT_SPECS, catalog_payload
from app.services.org_hierarchy import attach_child

BACKEND_APP = Path(__file__).resolve().parents[1] / 'app'


async def _with_session(coro_factory):
    engine = create_async_engine(settings.DATABASE_URL)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        async with session_factory() as db:
            result = await coro_factory(db)
            await db.commit()
            return result
    finally:
        await engine.dispose()


async def _make_org(db: AsyncSession, *, suffix: str) -> Organization:
    org = Organization(
        id=uuid4(),
        name=f'HR test {suffix}',
        slug=f'hr-{suffix}',
        plan='trial',
        is_active=True,
        settings={},
    )
    db.add(org)
    await db.flush()
    return org


async def _cleanup(db: AsyncSession, org_ids: list[UUID]) -> None:
    if not org_ids:
        return
    await db.execute(
        delete(OrgHierarchyLink).where(
            or_(
                OrgHierarchyLink.head_org_id.in_(org_ids),
                OrgHierarchyLink.child_org_id.in_(org_ids),
            )
        )
    )
    await db.execute(delete(Organization).where(Organization.id.in_(org_ids)))


def test_holding_reports_catalog_whitelist() -> None:
    catalog = {item['report_id']: item for item in catalog_payload()}
    assert 'shipments' in catalog
    assert 'group' in catalog['shipments']['modes']
    assert 'group' not in catalog['timesheet']['modes']
    assert catalog['salary']['group_unsupported_reason']
    assert 'marketplace' not in catalog


def test_tenant_reports_router_still_single_org() -> None:
    tree = ast.parse((BACKEND_APP / 'routers' / 'reports.py').read_text(encoding='utf-8'))
    names: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                names.add(alias.name)
        elif isinstance(node, ast.ImportFrom) and node.module:
            names.add(node.module)
    assert not any('holding' in n for n in names)


def test_holding_reports_service_no_marketplace() -> None:
    tree = ast.parse(
        (BACKEND_APP / 'services' / 'holding_reports.py').read_text(encoding='utf-8')
    )
    names: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                names.add(alias.name)
        elif isinstance(node, ast.ImportFrom) and node.module:
            names.add(node.module)
    assert not any('marketplace' in n or 'market_' in n for n in names)
    # Must call existing builders, not a second engine
    src = (BACKEND_APP / 'services' / 'holding_reports.py').read_text(encoding='utf-8')
    assert 'build_shipments_workbook' in src
    assert 'build_expenses_workbook' in src


def test_holding_reports_forbidden_without_links(
    client: httpx.Client,
    admin_headers: dict[str, str],
) -> None:
    catalog = client.get('/api/holding/reports/catalog', headers=admin_headers)
    assert catalog.status_code == 403, catalog.text

    today = date.today()
    export = client.post(
        '/api/holding/reports/export',
        headers=admin_headers,
        json={
            'report_id': 'shipments',
            'mode': 'group',
            'from_date': (today - timedelta(days=7)).isoformat(),
            'to_date': today.isoformat(),
        },
    )
    assert export.status_code == 403, export.text

    # Ordinary tenant report still works
    plain = client.post(
        '/api/reports/shipments',
        headers=admin_headers,
        json={
            'from_date': (today - timedelta(days=7)).isoformat(),
            'to_date': today.isoformat(),
        },
    )
    assert plain.status_code == 200, plain.text


def test_holding_report_whitelist_matches_frontend() -> None:
    """FE HOLDING_REPORT_SUPPORT keys/modes must match HOLDING_REPORT_SPECS."""
    import re

    fe_path = (
        Path(__file__).resolve().parents[2]
        / 'src'
        / 'features'
        / 'reports'
        / 'holdingSupport.ts'
    )
    text = fe_path.read_text(encoding='utf-8')
    # entries like: shipments: { modes: ['child', 'group'] }
    fe_modes: dict[str, set[str]] = {}
    for match in re.finditer(
        r"[\"']?([a-z0-9\-]+)[\"']?:\s*\{\s*modes:\s*\[([^\]]+)\]",
        text,
    ):
        rid = match.group(1)
        modes = {m.strip().strip("'\"") for m in match.group(2).split(',') if m.strip()}
        fe_modes[rid] = modes

    assert set(fe_modes) == set(HOLDING_REPORT_SPECS)
    for report_id, spec in HOLDING_REPORT_SPECS.items():
        assert fe_modes[report_id] == set(spec.modes), report_id
    assert 'marketplace' not in fe_modes


def test_holding_reports_child_and_group_export(
    client: httpx.Client,
    demo_org_id: str,
    admin_headers: dict[str, str],
) -> None:
    created: list[UUID] = []
    head_id = UUID(demo_org_id)

    async def seed(db: AsyncSession) -> str:
        child = await _make_org(db, suffix=uuid4().hex[:8])
        created.append(child.id)
        await attach_child(db, head_org_id=head_id, child_org_id=child.id)
        return str(child.id)

    async def cleanup(db: AsyncSession) -> None:
        await _cleanup(db, created)

    child_id = asyncio.run(_with_session(seed))
    today = date.today()
    period = {
        'from_date': (today - timedelta(days=30)).isoformat(),
        'to_date': today.isoformat(),
    }
    try:
        catalog = client.get('/api/holding/reports/catalog', headers=admin_headers)
        assert catalog.status_code == 200, catalog.text
        ids = {item['report_id'] for item in catalog.json()}
        assert ids == set(HOLDING_REPORT_SPECS.keys())

        child_export = client.post(
            '/api/holding/reports/export',
            headers=admin_headers,
            json={
                'report_id': 'expenses',
                'mode': 'child',
                'child_org_id': child_id,
                **period,
            },
        )
        assert child_export.status_code == 200, child_export.text
        assert 'spreadsheet' in child_export.headers.get('content-type', '')
        wb = openpyxl.load_workbook(BytesIO(child_export.content))
        assert 'Область' in wb.sheetnames
        assert wb['Область']['B2'].value == 'Отчёт по одной КФХ (holding)'

        group_export = client.post(
            '/api/holding/reports/export',
            headers=admin_headers,
            json={'report_id': 'expenses', 'mode': 'group', **period},
        )
        assert group_export.status_code == 200, group_export.text
        gwb = openpyxl.load_workbook(BytesIO(group_export.content))
        assert 'Область' in gwb.sheetnames
        assert 'Сводка по КФХ' in str(gwb['Область']['B2'].value)

        # timesheet group must be rejected
        bad = client.post(
            '/api/holding/reports/export',
            headers=admin_headers,
            json={'report_id': 'timesheet', 'mode': 'group', **period},
        )
        assert bad.status_code == 400, bad.text

        # foreign child_org_id forbidden
        foreign = client.post(
            '/api/holding/reports/export',
            headers=admin_headers,
            json={
                'report_id': 'expenses',
                'mode': 'child',
                'child_org_id': str(uuid4()),
                **period,
            },
        )
        assert foreign.status_code == 403, foreign.text

        # Single-org path unchanged
        plain = client.post(
            '/api/reports/expenses',
            headers=admin_headers,
            json=period,
        )
        assert plain.status_code == 200, plain.text
    finally:
        asyncio.run(_with_session(cleanup))


def test_holding_reports_manager_without_action_forbidden(
    client: httpx.Client,
    manager_headers: dict[str, str],
) -> None:
    r = client.get('/api/holding/reports/catalog', headers=manager_headers)
    assert r.status_code == 403, r.text
