"""Shipment requests Excel report + org feature flag + shift_id on complete."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from uuid import uuid4

import httpx
import pytest
from openpyxl import load_workbook


def _create_item(client: httpx.Client, headers: dict[str, str], *, stock: float = 100) -> str:
    response = client.post(
        '/api/inventory',
        headers=headers,
        json={
            'name': f'Report SR {uuid4().hex[:8]}',
            'category': 'other',
            'unit': 'кг',
            'current_stock': stock,
            'min_stock': 0,
            'total_capacity': 10000,
        },
    )
    assert response.status_code == 201, response.text
    return str(response.json()['id'])


def _create_request(
    client: httpx.Client,
    headers: dict[str, str],
    *,
    item_id: str,
    quantity: float = 5,
) -> dict:
    planned = (datetime.now(timezone.utc) + timedelta(days=1)).isoformat()
    response = client.post(
        '/api/shipment-requests',
        headers=headers,
        json={
            'customer_name': 'ООО Отчёт',
            'inventory_item_id': item_id,
            'quantity': quantity,
            'price': 20,
            'planned_at': planned,
            'priority': 'urgent',
        },
    )
    assert response.status_code == 201, response.text
    return response.json()


def test_shipment_requests_report_excel(
    client: httpx.Client,
    admin_headers: dict[str, str],
) -> None:
    item_id = _create_item(client, admin_headers)
    row = _create_request(client, admin_headers, item_id=item_id)

    today = datetime.now(timezone.utc).date()
    response = client.post(
        '/api/reports/shipment-requests',
        headers=admin_headers,
        json={'from_date': str(today - timedelta(days=1)), 'to_date': str(today)},
    )
    assert response.status_code == 200, response.text
    assert (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        in response.headers.get('content-type', '')
    )

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert 'Дата создания' in headers
    assert 'Контрагент' in headers
    assert 'ТМЦ' in headers
    assert 'Исполнитель' in headers
    assert 'Срочная' in headers
    assert 'ID смены' in headers

    values = [[cell.value for cell in excel_row] for excel_row in ws.iter_rows(min_row=2)]
    flat = ' '.join(str(v) for row_vals in values for v in row_vals if v is not None)
    assert 'ООО Отчёт' in flat
    assert row['id']  # created successfully; row may appear in sheet


def test_feature_flag_disables_shipment_requests_api(
    client: httpx.Client,
    admin_headers: dict[str, str],
) -> None:
    disabled = client.patch(
        '/api/settings/organization',
        headers=admin_headers,
        json={'shipment_requests_enabled': False},
    )
    assert disabled.status_code == 200, disabled.text
    assert disabled.json()['shipment_requests_enabled'] is False

    try:
        listed = client.get('/api/shipment-requests', headers=admin_headers)
        assert listed.status_code == 403, listed.text
    finally:
        enabled = client.patch(
            '/api/settings/organization',
            headers=admin_headers,
            json={'shipment_requests_enabled': True},
        )
        assert enabled.status_code == 200, enabled.text
        assert enabled.json()['shipment_requests_enabled'] is True

    ok = client.get('/api/shipment-requests', headers=admin_headers)
    assert ok.status_code == 200, ok.text


def test_complete_sets_shift_id_when_open_shift_exists(
    client: httpx.Client,
    admin_headers: dict[str, str],
    demo_org_id: str,
) -> None:
    item_id = _create_item(client, admin_headers, stock=50)
    row = _create_request(client, admin_headers, item_id=item_id, quantity=3)
    assert (
        client.post(f"/api/shipment-requests/{row['id']}/start", headers=admin_headers).status_code
        == 200
    )

    # Use admin as executor: open a shift for the current admin employee.
    me = client.get('/api/auth/me', headers=admin_headers)
    assert me.status_code == 200, me.text
    employee_id = me.json()['id']

    locs = client.get('/api/locations', headers=admin_headers).json()
    wts = client.get('/api/work-types', headers=admin_headers).json()
    loc = next(l for l in locs if l.get('is_active'))
    wt = next(w for w in wts if w.get('is_active'))

    # Close leftover open shifts for this employee.
    opens = client.get(
        '/api/shifts',
        headers=admin_headers,
        params={'status': 'open', 'employee_id': employee_id},
    )
    assert opens.status_code == 200, opens.text
    for shift in opens.json():
        client.post(
            f"/api/shifts/{shift['id']}/close",
            headers=admin_headers,
            json={'description': 'cleanup before shipment request shift_id test'},
        )

    opened = client.post(
        '/api/shifts',
        headers=admin_headers,
        json={
            'location_id': loc['id'],
            'work_type_id': wt['id'],
            'employee_id': employee_id,
        },
    )
    assert opened.status_code in (200, 201), opened.text
    shift_id = opened.json()['id']

    try:
        done = client.post(
            f"/api/shipment-requests/{row['id']}/complete",
            headers=admin_headers,
            json={},
        )
        assert done.status_code == 200, done.text
        assert done.json()['shift_id'] == shift_id
    finally:
        client.post(
            f'/api/shifts/{shift_id}/close',
            headers=admin_headers,
            json={'description': 'cleanup after shipment request shift_id test'},
        )


def test_complete_keeps_shift_id_null_without_open_shift(
    client: httpx.Client,
    admin_headers: dict[str, str],
) -> None:
    item_id = _create_item(client, admin_headers, stock=50)
    row = _create_request(client, admin_headers, item_id=item_id, quantity=2)
    assert (
        client.post(f"/api/shipment-requests/{row['id']}/start", headers=admin_headers).status_code
        == 200
    )

    me = client.get('/api/auth/me', headers=admin_headers)
    employee_id = me.json()['id']
    opens = client.get(
        '/api/shifts',
        headers=admin_headers,
        params={'status': 'open', 'employee_id': employee_id},
    ).json()
    for shift in opens:
        client.post(
            f"/api/shifts/{shift['id']}/close",
            headers=admin_headers,
            json={'description': 'ensure no open shift'},
        )

    done = client.post(
        f"/api/shipment-requests/{row['id']}/complete",
        headers=admin_headers,
        json={},
    )
    assert done.status_code == 200, done.text
    assert done.json()['shift_id'] is None
