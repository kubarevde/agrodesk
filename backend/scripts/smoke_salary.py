"""Smoke checks for employee rates + salary reports (org-scoped auth)."""

from __future__ import annotations

import io
import os
from datetime import date

import httpx
from openpyxl import load_workbook

BASE = os.environ.get('API_BASE_URL', 'http://localhost:8000')


def resolve_org_id(client: httpx.Client) -> str:
    env_org = os.environ.get('ORG_ID', '').strip()
    if env_org:
        return env_org
    r = client.get(f'{BASE}/api/auth/orgs')
    r.raise_for_status()
    orgs = r.json()
    assert orgs, 'No active organizations — run seed'
    return orgs[0]['id']


def login(client: httpx.Client, code: str, org_id: str) -> str:
    r = client.post(
        f'{BASE}/api/auth/login',
        json={'email': code, 'password': '1234', 'org_id': org_id},
    )
    r.raise_for_status()
    return r.json()['access_token']


def auth(token: str) -> dict[str, str]:
    return {'Authorization': f'Bearer {token}'}


def main() -> None:
    with httpx.Client(timeout=60) as c:
        org_id = resolve_org_id(c)
        print('org_id', org_id)

        admin = login(c, 'EMP000', org_id)
        mgr = login(c, 'EMP003', org_id)
        emp = login(c, 'EMP001', org_id)
        ha, hm, he = auth(admin), auth(mgr), auth(emp)

        emp1 = c.get(f'{BASE}/api/employees/me', headers=he).json()
        print('emp1', emp1['id'], emp1.get('full_name'))

        wts = c.get(f'{BASE}/api/work-types', headers=hm).json()
        field = next(
            (
                w
                for w in wts
                if 'полев' in (w.get('category') or '').lower()
                or 'полев' in w['name'].lower()
            ),
            None,
        ) or next((w for w in wts if w['name'] == 'Посев'), None) or wts[0]
        print('wt', field['name'], field.get('category'), field['id'])

        opens = c.get(
            f'{BASE}/api/shifts',
            headers=hm,
            params={'status': 'open', 'employee_id': emp1['id']},
        ).json()
        for s in opens:
            rr = c.post(
                f"{BASE}/api/shifts/{s['id']}/close",
                headers=he,
                json={'description': 'cleanup'},
            )
            print('close open', s['id'], rr.status_code)

        today = date.today().isoformat()
        rates = c.get(
            f'{BASE}/api/employee-rates',
            headers=hm,
            params={'employee_id': emp1['id']},
        ).json()
        print('existing rates', len(rates) if isinstance(rates, list) else rates)
        if isinstance(rates, list):
            for rate in rates:
                if rate.get('work_type_id') == field['id'] and rate.get('valid_to') is None:
                    d = c.delete(f"{BASE}/api/employee-rates/{rate['id']}", headers=ha)
                    print('deleted rate', rate['id'], d.status_code)

        body = {
            'employee_id': emp1['id'],
            'work_type_id': field['id'],
            'rate': 500,
            'overtime_multiplier': 1.33,
            'overtime_threshold_hours': 8.0,
            'valid_from': today,
            'notes': 'smoke rate',
        }
        pr = c.post(f'{BASE}/api/employee-rates', headers=ha, json=body)
        print('POST rate', pr.status_code, pr.text[:300])
        assert pr.status_code == 201, pr.text

        prev = c.get(
            f'{BASE}/api/employee-rates/calculate-preview',
            headers=hm,
            params={
                'employee_id': emp1['id'],
                'work_type_id': field['id'],
                'hours': 10,
                'shift_date': today,
            },
        )
        print('preview', prev.status_code, prev.text)

        locs = c.get(f'{BASE}/api/locations', headers=hm).json()
        loc = next(l for l in locs if l.get('is_active') and 'Поле' in l['name'])
        eq_list = c.get(f'{BASE}/api/equipment', headers=hm).json()
        eq = next(e for e in eq_list if e.get('is_active') and e['name'] == 'МТЗ-82')
        manual = {
            'employee_id': emp1['id'],
            'date': today,
            'start_time': '08:00:00',
            'end_time': '18:00:00',
            'location_id': loc['id'],
            'work_type_id': field['id'],
            'equipment_id': eq['id'],
            'description': 'salary smoke 10h',
        }
        mr = c.post(f'{BASE}/api/shifts/manual', headers=hm, json=manual)
        print('manual', mr.status_code, mr.text[:500])
        assert mr.status_code in (200, 201), mr.text
        sh = mr.json()
        print('calculated_amount', sh.get('calculated_amount'), 'snapshot', sh.get('rate_snapshot'))
        assert float(sh['calculated_amount']) == 5330.0, sh

        month = today[:7]
        sr = c.post(f'{BASE}/api/reports/salary', headers=hm, json={'month': month})
        print('salary xlsx', sr.status_code, sr.headers.get('content-type'), len(sr.content))
        assert sr.status_code == 200, sr.text[:300]
        wb = load_workbook(io.BytesIO(sr.content))
        print('sheets', wb.sheetnames)
        # Sheet titles: Сводка / По сменам / Ставки
        assert len(wb.sheetnames) == 3, wb.sheetnames

        prevj = c.get(f'{BASE}/api/reports/salary-preview', headers=hm, params={'month': month})
        print(
            'preview json',
            prevj.status_code,
            list(prevj.json().keys()) if prevj.status_code == 200 else prevj.text[:200],
        )
        earn = c.get(f'{BASE}/api/employees/me/earnings', headers=he, params={'month': month})
        print(
            'earnings',
            earn.status_code,
            list(earn.json().keys()) if earn.status_code == 200 else earn.text[:200],
        )
        print('ALL SMOKE OK')


if __name__ == '__main__':
    main()
